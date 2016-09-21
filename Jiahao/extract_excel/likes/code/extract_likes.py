from openpyxl import load_workbook
from openpyxl import Workbook

inputFile = "../input/profile_likes_excel.xlsx"

wb = load_workbook(inputFile)
ws = wb.get_sheet_by_name('Sheet1')

class Person:
    def __init__(self, id = '', name = '', likes = []):
    	self.id = id
        self.name = name
        self.likes = likes

    def show(self):
    	r = [self.id, self.name]
    	for i in self.likes:
    		r += [i]
    	return r

def get_info(ws):
	users = []
	maxl = 0
	for row in range(2, 83):
		i = ws['C' + str(row)].value.split('/')[3]
		likes = ws['D' + str(row)].value
		name = get_name(likes)
		l = clean_likes(likes)
		if maxl < len(l):
			maxl = len(l)
		current_user = Person(i, name, l)
		users += [current_user]
	return users, maxl

def get_name(s):
	count = 0
	flag = 0
	for c in s:
		if c in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
			flag += 1
			if flag == 3:
				return s[0:count]
		count += 1

def clean_likes(s):
	r = []
	for z in s.split('Like')[3:]:
		if len(z) < 50:
			r += [z]
	return r

def generate_output(persons, maxl):
	wb = Workbook(write_only=True)
	ws = wb.create_sheet()
	first = ['id', 'name']
	for i in range(1, maxl+1):
		first += ['like_' + str(i)]
	ws.append(first)
	for p in persons:
		ws.append(p.show())
	wb.save('../output/people_likes_output.xlsx') 

persons, maxl = get_info(ws)
print len(persons)
generate_output(persons, maxl)











