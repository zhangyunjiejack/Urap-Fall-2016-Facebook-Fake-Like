from openpyxl import load_workbook
from openpyxl import Workbook

inputFile = "../input/profile_friends_excel.xlsx"

wb = load_workbook(inputFile)
ws = wb.get_sheet_by_name('Sheet1')

class Person:
    def __init__(self, id = '', name = '', friends = []):
    	self.id = id
        self.name = name
        self.friends = friends
        self.friendsNumber = len(friends)

    def show(self):
    	r = [self.id, self.name, self.friendsNumber]
    	for i in self.friends:
    		r += [i]
    	return r

def get_info(ws):
	users = []
	maxl = 0
	for row in range(2, 83):
		i = ws['C' + str(row)].value.split('/')[3]
		friends = ws['E' + str(row)].value
		name = get_name(friends)
		l = clean_friends(friends)
		if maxl < len(l):
			maxl = len(l)
		current_user = Person(i, name, l)
		users += [current_user]
	return users, maxl

def get_name(s):
	count = 0
	flag = 0
	for c in s:
		if c in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
			flag += 1
			if flag == 3:
				return s[0:count]
		count += 1

def clean_friends(s):
	r = []
	for z in s.split('Add FriendFriend Request Sent')[3:]:
		if len(z) < 50:
			r += [z]
	return r

def generate_output(persons, maxl):
	wb = Workbook(write_only=True)
	ws = wb.create_sheet()
	first = ['id', 'name', 'friendsNumber']
	for i in range(1, maxl+1):
		first += ['friend_' + str(i)]
	ws.append(first)
	for p in persons:
		ws.append(p.show())
	wb.save('../output/people_friends_output.xlsx') 

persons, maxl = get_info(ws)
print len(persons)
generate_output(persons, maxl)











